from toolboxTMU import parameter, sqlLibrary, find_tap, initParameter, dataParser, harmonicParser, convertBinList
from openpyxl import Workbook
import requests
import mysql.connector, time, datetime, math, openpyxl, sys, shutil, os
from requests.models import StreamConsumedError
from requests.exceptions import Timeout
import random
import RPi.GPIO as GPIO

engineName = "Trafo X"
teleURL = 'http://192.168.4.120:1444/api/transformer/sendNotificationToTelegramGroup'
progStat = True
debugMsg = False
infoMsg = True
dryType = False
gasType = False
transmitterModeMinus = False

exhibitStat = False
OLTCstat = False
pressureStat = True
tempStat = True

GPIO.setmode(GPIO.BCM)
GPIO.setup(13, GPIO.IN)
GPIO.setup(17, GPIO.IN)

source = {
    16: 862,
    15: 817,
    14: 773,
    13: 728,
    12: 683,
    11: 639,
    10: 594,
    9: 548,
    8: 503,
    7: 458,
    6: 413,
    5: 367,
    4: 321,
    3: 276,
    2: 230,
    1: 184,
    0: 0
}

def main():
    if infoMsg == True: print("1D|Initialize Program") 
    dataLen = 56
    watchedData = 29
    cycleTime = 2 / 60
    
    db = mysql.connector.connect(
        host = "localhost",
        user = "client",
        passwd = "raspi",
        database= "iot_trafo_client")
    cursor = db.cursor()

    #init logger rawdata
    ts = time.strftime("%Y%m%d")
    pathStr = r'/home/pi/tmu/tmu-app-client-deploy/assets/datalog/rawdata/datalogger-'
    #pathStr = r'/home/pi/tmu-v2-smart/assets/rawdata-test/datalogger-'
    pathDatLog = pathStr + ts + '.xlsx'
    pathBkup = r'/home/pi/tmu-v2-smart/assets/rawdata-test/backup/datalogger-backup-'
    pathDatBkup = pathBkup + ts + '.xlsx'
     
    try:
        wb = openpyxl.load_workbook(pathDatLog)
        if infoMsg == True: print("1D|Open Existing Excel")
    except:
        #create new datalog
        workbook = Workbook()
        workbook.save(pathDatLog)
        #create datalog's header
        wb = openpyxl.load_workbook(pathDatLog)
        sheet = wb.active
        sheet.title = "Raw_data"
        name = (('timestamp', 
                    'V-un', 'V-vn', 'V-wn', 'V-uv', 'V-vw', 'V-uw',
                    'I-u', 'I-v', 'I-w', 'Iavg', 'In',
                    'THDV-u', 'THDV-v', 'THDV-w', 'THDI-u', 'THDI-v', 'THDI-w',
                    'P-u', 'P-v', 'P-w', 'Ptot',
                    'Q-u', 'Q-v', 'Q-w', 'Q-tot',
                    'S-u', 'S-v', 'S-w', 'S-tot',
                    'PF-u', 'PF-v', 'PF-w', 'PFavg', 'Freq', 'kWh', 'kVARh',
                    'BusTemp-u', 'BusTemp-v', 'BusTemp-w', 'OilTemp',
                    'WTITemp-u', 'WTITemp-v', 'WTITemp-w',  'Press', 'Level',
                    'KRated-u', 'Derating-u', 'KRated-v', 'Derating-v', 'KRated-w', 'Derating-w',
                    'H2ppm', 'Moistppm', 'Vdiff-uv', 'Vdiff-vw', 'Vdiff-uw',
                    'trafoStatus', 'DIstat', 'DOstat', 'Alarm', 'Trip1', 'Trip2', 'Tap Position'),)
        for row in name:
            sheet.append(row)
        wb.save(pathDatLog)
        if debugMsg == True: print("1D|Create New Excel")
    
    inputData = [0]*dataLen
    currentStat = [0]*watchedData
    currentTrip = [0]*watchedData
    dataName = ['']*watchedData
    activeParam = [None]*watchedData
    activeFailure = [None]*watchedData
    dataSet = [parameter("Name", 0, False, None, None, None, None, 3, 0)]
    for i in range(0, dataLen-1):
        dataSet.append(parameter("Name", 0, False, None, None, None, None, 3, 0))
    messageReason = ['Extreme Low',
                'Low', 
                'Back Normal', 
                'High', 
                'Extreme High']
    msgEvent = [None] * watchedData
    msgReminder = [None] * watchedData
    telePrevTime = excelPrevTime  = excelSavePrevTime = datetime.datetime.now()
    cursor.execute(sqlLibrary.sqlFailure)
    listFailure = cursor.fetchall()
    for i in range(0, len(listFailure)):
        if listFailure[i][2] == None:
            activeFailure[activeFailure.index(None)] = listFailure[i]
    
    if infoMsg == True: print("1D|Start Loop")
    while progStat:
        if debugMsg == True: print("1D|1 Fetch DB Data")
        start_time = time.time()
        cursor.execute(sqlLibrary.sqlTrafoSetting)
        trafoSetting = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlTrafoData)
        trafoData = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlTripSetting)
        tripSetting = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlDIscan)
        inputIO = cursor.fetchall()
        cursor.execute(sqlLibrary.sqlDOscan)
        outputIO = cursor.fetchall()
        cursor.execute(sqlLibrary.sqlConstantWTI, (str(trafoData[27]), ))
        constantWTI = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlTrafoStatus)
        prevStat = list(cursor.fetchall()[0][1:])
        cursor.execute(sqlLibrary.sqlTripStatus)
        prevTrip = list(cursor.fetchall()[0][1:])
        db.commit()

        oilLevelAlarm = GPIO.input(13)
        oilLevelTrip = GPIO.input(17)
        analogIn1 = inputIO[6][2]
        analogIn2 = inputIO[7][2]

        if (oilLevelAlarm and oilLevelTrip) or oilLevelTrip:
            oilStat = 1
        elif oilLevelAlarm:
            oilStat = 2
        elif oilLevelAlarm == 0 and oilLevelTrip == 0:
            oilStat = 3
        inputData[44] = oilStat     #Oil Level

        if tempStat :
            if transmitterModeMinus :
                suhu = round(((analogIn1 * 0.009537) - 112.5), 3)
            else:
                suhu = round(((analogIn1 * 0.007630) - 50), 3)
            inputData[39] = max(0, suhu)
        else : 
            inputData[39] = 0

        if pressureStat:
            inputData[43] = (analogIn2 - 6553) / 26214
        else:
            inputData[43] = 0
        
        if debugMsg == True: print("1D|9 Input all data DB")
        dataResult = initParameter(dataSet, inputData, trafoSetting, trafoData, tripSetting, dataLen) 
        sendData = [datetime.datetime.now()] + inputData
        cursor.execute(sqlLibrary.sqlInsertData, sendData)
        db.commit()
        if debugMsg == True: print("1D|10 Check Failures Stat")
        maxStat = 0
        i =  0
        for data in dataResult:
            if data.isWatched:
                maxStat = data.trafoStat if data.trafoStat > maxStat else maxStat
                currentStat[i] = data.status
                currentTrip[i] = data.trafoStat
                dataName[i] = data.name
                #print(data.name)
                if data.status != prevStat[i]:
                    if data.status != 3:
                        if data.name in activeParam:
                            lastTimestamp = activeFailure[activeParam.index(data.name)][1]
                            duration = int((datetime.datetime.now() - lastTimestamp).total_seconds())
                            errorVal = [duration, activeFailure[activeParam.index(data.name)][0]]
                            cursor.execute(sqlLibrary.sqlResolveFailure, errorVal)
                            activeFailure[activeParam.index(data.name)] = None
                            activeParam[activeParam.index(data.name)] = None
                        errorVal = [datetime.datetime.now(), messageReason[data.status - 1], data.name, str(data.value)]
                        cursor.execute(sqlLibrary.sqlInsertFailure, errorVal)
                        cursor.execute(sqlLibrary.sqlLastFailure)
                        lastActive = cursor.fetchall()[0]
                        activeFailure[activeFailure.index(None)] = lastActive
                        loadProfile = str((round((data.value / trafoData[6]) * 10000))/100) + " Percent , Rated Current = " + str(trafoData[6])
                        msgEvent[i] = str(data.name + " " + messageReason[data.status - 1] + " , Value = " + (loadProfile if i == 3 or i == 4 or i == 5 else str(data.value)) + "\n" + "Time Occurence : " + str(datetime.datetime.now()))
                    elif data.status == 3:
                        lastTimestamp = activeFailure[activeParam.index(data.name)][1]
                        duration = int((datetime.datetime.now() - lastTimestamp).total_seconds())
                        errorVal = [duration, activeFailure[activeParam.index(data.name)][0]]
                        cursor.execute(sqlLibrary.sqlResolveFailure, errorVal)
                        activeFailure[activeParam.index(data.name)] = None
                        activeParam[activeParam.index(data.name)] = None
                        msgEvent[i] = None
                i = i + 1
        if debugMsg == True: print("1D|11 Check state changes")
        if prevStat != currentStat or prevTrip != currentTrip:
            #print("Send Telegram Lhooo")
            tele = list(filter(None, msgEvent))
            if tele:
                for message in tele:                
                    messages = engineName + " Says : " + "\n" + message
                    pload = {'message':messages}
                    try:
                        r = requests.post(teleURL, data = pload, timeout = 5, verify = False)
                    except Timeout:
                        pass
                    except Exception as Argument:
                        pass
            else:
                pass
            cursor.execute(sqlLibrary.sqlUpdateTransformerStatus, currentStat)
            cursor.execute(sqlLibrary.sqlUpdateTripStatus, currentTrip)
            cursor.execute(sqlLibrary.sqlUpdateTrafoStat, (maxStat,))
            db.commit()
        else:
            pass
        binList = convertBinList(inputIO, outputIO, currentTrip)
        if int((datetime.datetime.now() - telePrevTime).total_seconds()) > 3600:
            if debugMsg == True: print("1D|12 Routine remind Tele")
            #print("sekadar mengingatkan")
            for i in range(0, len(activeFailure)):
                if activeFailure[i]:
                    failureIndex = dataName.index(activeFailure[i][4])
                    msgReminder[failureIndex] = str(activeFailure[i][4] + " " + activeFailure[i][3] + " , Value = " + activeFailure[i][5] + "\n" + "Time Occurence : " + str(activeFailure[i][1]))                    
                    messages = engineName + " Says : " + "\n" + msgReminder
                    pload = {'message':messages}
                    try:
                        r = requests.post(teleURL, data = pload, timeout = 5, verify = False)
                    except Timeout:
                        pass
                    except Exception as Argument:
                        pass

            telePrevTime = datetime.datetime.now()
        #print(inputData)
        if int((datetime.datetime.now() - excelPrevTime).total_seconds()) > 3:
            if debugMsg == True: print("1D|12A Routine Add data to work stage excel")
            sendLog = [datetime.datetime.now().strftime("%H:%M:%S")] + inputData + [maxStat] + binList + [OLTCstat]
            sendLog = ((tuple(sendLog)),)
            sheet = wb["Raw_data"]
            for row in sendLog:
                sheet.append(row)
            excelPrevTime = datetime.datetime.now()
        if int((datetime.datetime.now() - excelSavePrevTime).total_seconds()) > 180:
            if debugMsg == True: print("1D|12B Routine Save Excel")
            if infoMsg == True: print("1D|Check Current Excel Size")
            if os.path.isfile(pathDatBkup) and os.path.getsize(pathDatBkup) >= os.path.getsize(pathDatLog):
                if infoMsg == True: print("1D|Excel Smaller than backup, replacing")
                shutil.copy2(pathDatBkup, pathDatLog)
            else:
                #create backup
                if infoMsg == True: print("1D|Backup Excel")
                shutil.copy2(pathDatLog, pathDatBkup)
            #print("save excel data here")
            try:
                if infoMsg == True: print("1D|Try to save Excel from work stage")
                wb.save(pathDatLog)
                time.sleep(0.5)
                if infoMsg == True: print("1D|Excel Size %s " % (os.path.getsize(pathDatLog)))
                if infoMsg == True: print("1D|Backup Size %s " % (os.path.getsize(pathDatBkup)))
                if (os.path.getsize(pathDatBkup) - os.path.getsize(pathDatLog)) < 3000:
                    if infoMsg == True: print("1D|Save Success")
                else:
                    raise Exception("backup larger than saved excel")
            except Exception as e:
                if infoMsg == True: print("1D|%s" % e)
                if infoMsg == True: print("1D|Save Failed, return to backup, restart system")
                shutil.copy2(pathDatBkup, pathDatLog)
                if infoMsg == True: print("1D|Restart")
            excelSavePrevTime = datetime.datetime.now()
                        
        cycleTime = (round(10000 * (time.time() - start_time)))/10000
        if debugMsg == True: print("1D|Cycle time %s" % cycleTime)
        print("1T|%s" % datetime.datetime.now())
        sys.stdout.flush()
        time.sleep(4)
        
if __name__ == "__main__":
    main()
